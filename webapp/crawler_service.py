import sys
import os
import shutil
import tempfile
import logging
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

logger = logging.getLogger(__name__)


def run_crawler(keywords: list, from_date: str, to_date: str, output_dir: str):
    """
    Run G2B crawler and save result Excel to output_dir.
    Returns (output_filepath, row_count) or (None, 0) on failure.
    Each call uses an isolated temp directory so concurrent runs don't interfere.
    """
    from web_crawler import get_session, download_excel_by_keyword, merge_excel_files
    from openpyxl import load_workbook

    os.makedirs(output_dir, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmpdir:
        excel_files = []
        for keyword in keywords:
            session = get_session()
            if session is None:
                logger.warning(f"'{keyword}' 세션 발급 실패, 건너뜀")
                continue
            excel_file = download_excel_by_keyword(
                session, keyword, from_date, to_date, work_dir=tmpdir
            )
            excel_files.append(excel_file)

        valid_files = [f for f in excel_files if f is not None]
        if not valid_files:
            logger.error("다운로드된 파일 없음")
            return None, 0

        tmp_output = merge_excel_files(
            excel_files, from_date, to_date, keywords, work_dir=tmpdir
        )
        if tmp_output is None:
            return None, 0

        dest_path = os.path.join(output_dir, os.path.basename(tmp_output))
        shutil.move(tmp_output, dest_path)

        try:
            wb = load_workbook(dest_path, read_only=True)
            ws = wb.active
            row_count = max(0, (ws.max_row or 5) - 5)
            wb.close()
        except Exception:
            row_count = 0

        return dest_path, row_count
