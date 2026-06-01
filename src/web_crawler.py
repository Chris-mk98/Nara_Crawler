import requests
import json
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
import copy

def get_session():
    """세션 생성 함수"""
    print("⏳ 세션 발급 중...")
    session = requests.Session()
    base_url = "https://www.g2b.go.kr"
    
    session_headers = {
        "Accept": "application/json",
        "Content-Type": "application/json;charset=UTF-8",
        "Referer": "https://www.g2b.go.kr/",
        "Origin": "https://www.g2b.go.kr",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    }
    
    session_response = session.post(f"{base_url}/co/coz/coza/util/getSession.do", headers=session_headers)
    if session_response.status_code == 200:
        print("✅ 세션 발급 성공")
        return session
    else:
        print("❌ 세션 발급 실패:", session_response.status_code)
        return None

def download_excel_by_keyword(session, keyword, from_date, to_date):
    """키워드별 엑셀 다운로드 함수"""
    print(f"⏳ '{keyword}' 키워드로 검색 중...")
    
    base_url = "https://www.g2b.go.kr"
    excel_url = f"{base_url}/pn/pnp/pnpe/BidPbac/selectBidPbacListExcel.do"
    
    payload = {
        "dlBidPbancLstM": {
            "untyBidPbancNo": "",
            "bidPbancNo": "",
            "bidPbancOrd": "",
            "prcmBsneUntyNoOrd": "",
            "prcmBsneSeCd": "0000 조070001 조070002 조070003 조070004 조070005 민079999",
            "bidPbancNm": keyword,
            "pbancPstgDt": "",
            "ldocNoVal": "",
            "bidPrspPrce": "",
            "ctrtDmndRcptNo": "",
            "dmstcOvrsSeCd": "",
            "pbancKndCd": "공440002",
            "ctrtTyCd": "",
            "bidCtrtMthdCd": "",
            "scsbdMthdCd": "",
            "fromBidDt": from_date,
            "toBidDt": to_date,
            "minBidPrspPrce": "",
            "maxBidPrspPrce": "",
            "bsneAllYn": "Y",
            "frcpYn": "Y",
            "rsrvYn": "Y",
            "laseYn": "Y",
            "untyGrpGb": "",
            "dmstNm": "",
            "pbancPicNm": "",
            "odnLmtLgdngCd": "",
            "odnLmtLgdngNm": "",
            "intpCd": "",
            "intpNm": "",
            "dtlsPrnmNo": "",
            "dtlsPrnmNm": "",
            "slprRcptDdlnYn": "",
            "lcrtTyCd": "",
            "isMas": "",
            "isElpdt": "",
            "oderInstUntyGrpNo": "",
            "instSearchRangeYn": "",
            "esdacYn": "",
            "infoSysCd": "정010029",
            "contxtSeCd": "콘010006",
            "bidDateType": "R",
            "brcoOrgnCd": "",
            "deptOrgnCd": "",
            "isShop": "",
            "srchTy": "0",
            "cangParmVal": "untySrch001",
            "currentPage": "",
            "recordCountPerPage": "10",
            "startIndex": 1,
            "endIndex": 10,
            "excelHeaderText": ""
        }
    }
    
    excel_headers = {
        "Accept": "*/*",
        "Content-Type": "application/json;charset=UTF-8",
        "Referer": "https://www.g2b.go.kr/",
        "Origin": "https://www.g2b.go.kr",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    }
    
    response = session.post(excel_url, headers=excel_headers, json=payload)
    
    if response.status_code == 200:
        temp_filename = f"temp_{keyword.replace(' ', '_')}.xlsx"
        with open(temp_filename, "wb") as f:
            f.write(response.content)
        print(f"✅ '{keyword}' 검색 완료")
        return temp_filename
    else:
        print(f"❌ '{keyword}' 검색 실패:", response.status_code)
        return None

def read_search_conditions():
    """검색 조건 파일 읽기"""
    print("⏳ 검색 조건 파일 읽는 중...")
    
    try:
        wb = load_workbook('nara_keyword_date.xlsx')
        ws = wb.active
        
        from_date = str(ws['C2'].value)
        to_date = str(ws['C3'].value)
        
        keywords = []
        row = 6
        while True:
            keyword = ws[f'B{row}'].value
            if keyword is None or str(keyword).strip() == '':
                break
            keywords.append(str(keyword).strip())
            row += 1
        
        print(f"✅ 검색 기간: {from_date} ~ {to_date}")
        print(f"✅ 검색 키워드 {len(keywords)}개: {keywords}")
        
        return from_date, to_date, keywords
    
    except Exception as e:
        print(f"❌ 검색 조건 파일 읽기 실패: {e}")
        return None, None, None

def merge_excel_files(excel_files, from_date, to_date, keywords):
    """엑셀 파일들을 병합하고 중복 제거"""
    print("⏳ 엑셀 파일 병합 및 중복 제거 중...")
    
    all_data = []
    meta_info = None
    headers = None
    original_wb = None
    
    for excel_file in excel_files:
        if excel_file is None:
            continue
            
        try:
            # 첫 번째 유효한 파일에서 원본 워크북 스타일 저장
            if original_wb is None:
                original_wb = load_workbook(excel_file)
                print("✅ 원본 파일 스타일 저장")
            
            # 전체 시트를 읽어서 메타정보와 헤더, 데이터 분리
            df_full = pd.read_excel(excel_file, header=None)
            
            # 메타정보 저장 (첫 번째 파일에서만)
            if meta_info is None:
                meta_info = df_full.iloc[0:3]
            
            # 헤더 저장 (첫 번째 파일에서만)
            if headers is None:
                headers = df_full.iloc[4:5]
            
            # 데이터 부분 (6행부터, 인덱스로는 5부터)
            data_part = df_full.iloc[5:]
            if not data_part.empty:
                all_data.append(data_part)
                
        except Exception as e:
            print(f"❌ {excel_file} 읽기 실패: {e}")
    
    if not all_data:
        print("❌ 병합할 데이터가 없습니다.")
        # 임시 파일 삭제
        for excel_file in excel_files:
            if excel_file and os.path.exists(excel_file):
                os.remove(excel_file)
        return None
    
    # 모든 데이터 병합
    combined_data = pd.concat(all_data, ignore_index=True)
    
    # D컬럼(입찰공고번호) 기준으로 중복 제거
    print(f"⏳ 중복 제거 전 데이터 행 수: {len(combined_data)}")
    combined_data = combined_data.drop_duplicates(subset=[combined_data.columns[3]], keep='first')
    # 인덱스 재정렬 (빈 행 제거)
    combined_data = combined_data.reset_index(drop=True)
    print(f"✅ 중복 제거 후 데이터 행 수: {len(combined_data)}")
    
    # 최종 엑셀 파일 생성
    save_time = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"나라장터_입찰공고_{from_date}_{to_date}_{save_time}.xlsx"
    
    # 새 워크북 생성
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = '검색결과'
    
    # 원본 워크시트에서 스타일 복사 (A1:H5)
    if original_wb:
        original_ws = original_wb.active
        print("⏳ 원본 스타일 복사 중...")
        
        # A1:H5 영역의 스타일 복사
        for row in range(1, 6):  # 1행부터 5행까지
            for col in range(1, 9):  # A열부터 H열까지
                original_cell = original_ws.cell(row=row, column=col)
                new_cell = new_ws.cell(row=row, column=col)
                
                # 값 복사
                if row <= 3:  # 메타정보 (A1:A3)
                    if row-1 < len(meta_info) and col-1 < len(meta_info.columns):
                        value = meta_info.iloc[row-1, col-1]
                        if pd.notna(value):
                            new_cell.value = value
                elif row == 5:  # 헤더 (A5:H5)
                    if col-1 < len(headers.columns):
                        value = headers.iloc[0, col-1]
                        if pd.notna(value):
                            new_cell.value = value
                
                # 스타일 복사
                if original_cell.has_style:
                    new_cell.font = copy.copy(original_cell.font)
                    new_cell.border = copy.copy(original_cell.border)
                    new_cell.fill = copy.copy(original_cell.fill)
                    new_cell.number_format = original_cell.number_format
                    new_cell.protection = copy.copy(original_cell.protection)
                    new_cell.alignment = copy.copy(original_cell.alignment)
    
    # 데이터 입력 및 테두리 적용 (6행부터)
    print("⏳ 데이터 입력 및 테두리 적용 중...")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    data_start_row = 6
    for i, (_, row_data) in enumerate(combined_data.iterrows()):
        current_row = i + data_start_row  # 순차적으로 6, 7, 8... 배치
        for j, value in enumerate(row_data):
            if j < 8:  # H열까지만
                cell = new_ws.cell(row=current_row, column=j+1)
                if pd.notna(value):
                    cell.value = value
                # 모든 데이터 셀에 테두리 적용
                cell.border = thin_border
    
    # 검색 조건 메모 시트 생성
    memo_ws = new_wb.create_sheet(title='검색조건')
    memo_data = [
        ['구분', '내용'],
        ['검색 시작일자', from_date],
        ['검색 종료일자', to_date],
        ['검색 키워드', ', '.join(keywords)]
    ]
    
    for row_idx, row_data in enumerate(memo_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            memo_ws.cell(row=row_idx, column=col_idx, value=value)
    
    # A:H 컬럼 자동 너비 조정
    print("⏳ 컬럼 너비 자동 조정 중...")
    for column in new_ws.columns:
        if column[0].column <= 8:  # A~H 컬럼만 (1~8)
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 최소 너비 8, 최대 너비 50으로 제한
            adjusted_width = min(max(max_length + 2, 8), 50)
            new_ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    new_wb.save(filename)
    
    # 임시 파일 삭제
    for excel_file in excel_files:
        if excel_file and os.path.exists(excel_file):
            os.remove(excel_file)
    
    print(f"✅ 최종 파일 저장 완료: {filename}")
    return filename

def main():
    """메인 실행 함수"""
    print("🚀 나라장터 입찰공고 검색 시작")
    
    # 1. 검색 조건 읽기
    from_date, to_date, keywords = read_search_conditions()
    if not keywords:
        print("❌ 검색 조건을 읽을 수 없습니다.")
        return
    
    # 2. 각 키워드별로 검색 및 다운로드
    excel_files = []
    for keyword in keywords:
        # 매번 새로운 세션 생성
        session = get_session()
        if session is None:
            continue
            
        excel_file = download_excel_by_keyword(session, keyword, from_date, to_date)
        excel_files.append(excel_file)
    
    # 3. 파일 병합 및 중복 제거
    final_file = merge_excel_files(excel_files, from_date, to_date, keywords)
    
    if final_file:
        print(f"🎉 모든 작업 완료! 결과 파일: {final_file}")
    else:
        print("❌ 작업 실패")

if __name__ == "__main__":
    main()